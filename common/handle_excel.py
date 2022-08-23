import openpyxl

class Handle_excel:
    def __init__(self,filename,sheetname):
        self.filename = filename
        self.sheetname = sheetname
    # 读取文件
    def Read_excel(self):
        workbook = openpyxl.load_workbook(self.filename)
        sheet = workbook[self.sheetname]
        res = list(sheet.rows)
        title = [t.value for t in res[0]]
        cases = []
        for i in res[1:]:
            data = [v.value for v in i]
            case = dict(zip(title,data))
            cases.append(case)
        return cases
        print(cases)
    # 写入文件
    def Write_excel(self,row,column,value):
        workbook = openpyxl.load_workbook(self.filename)
        sheet = workbook[self.sheetname]
        # 往哪一行写入数据
        sheet.cell(row=row,column=column,value=value)
        # 关闭文件
        workbook.save(filename=self.filename)




# if __name__ == '__main__':
#     a = Handle_excel(filename="/Users/taoqi/888/pythonProject5/data/test_data.xlsx",sheetname="token")
#     print(a.Read_excel())